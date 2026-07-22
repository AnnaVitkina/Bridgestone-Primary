from pathlib import Path
import re
import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
PROCESSING_DIR = BASE_DIR / "processing"
OUTPUT_DIR = BASE_DIR / "output"
ADDITION_DIR = BASE_DIR / "addition"
POSTAL_CODE_ZONES_FILE = "Postal Code Zones.txt"


def postal_code_zones_path() -> Path:
    return ADDITION_DIR / POSTAL_CODE_ZONES_FILE

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xlsb", ".csv"}


CONDITION_RULES = [
    {
        "Name": "Truck 115/120 M3",
        "Operator": "equals",
        "Values": "'22/Truck 115 M3', '05/Truck 120 M3'",
    },
    {
        "Name": "Dry 60 M3/H-Cube 76 M3",
        "Operator": "equals",
        "Values": "'09/TC 40\" Dry 60 M3', '10/TC 40\" H-Cube 76 M3'",
    },
    {
        "Name": "Truck 100/115 M3",
        "Operator": "equals",
        "Values": "'01/Truck 100 M3', '22/Truck 115 M3'",
    },
    {
        "Name": "Truck 100/95 M3",
        "Operator": "equals",
        "Values": "'01/Truck 100 M3', '21/Truck 95 M3'",
    },
    {
        "Name": "100 Cubic M",
        "Operator": "equals",
        "Values": "'01/Truck 100 M3'",
    },
    {
        "Name": "95 Cubic M",
        "Operator": "equals",
        "Values": "'21/Truck 95 M3'",
    },
    {
        "Name": "45ft Container",
        "Operator": "equals",
        "Values": "'12/TC 45\" Dry 88 M3'",
    },
    {
        "Name": "90 Cubic M",
        "Operator": "equals",
        "Values": "'19/Truck 90 M3'",
    },
    {
        "Name": "40ft HC Container",
        "Operator": "equals",
        "Values": "'10/TC 40\" H-Cube 76 M3'",
    },
    {
        "Name": "Truck 100/90 M3",
        "Operator": "equals",
        "Values": "'19/Truck 90 M3', '01/Truck 100 M3'",
    },
    {
        "Name": "40ft Container",
        "Operator": "equals",
        "Values": "'09/TC 40\" Dry 60 M3'",
    },
    {
        "Name": "45ft Container/40ft HC Container",
        "Operator": "equals",
        "Values": "'12/TC 45\" Dry 88 M3', '10/TC 40\" H-Cube 76 M3'",
    },
    {
        "Name": "40ft Container/40ft HC Container",
        "Operator": "equals",
        "Values": "'09/TC 40\" Dry 60 M3', '10/TC 40\" H-Cube 76 M3'",
    },
    {
        "Name": "Truck 100 M3/45 ft Container",
        "Operator": "equals",
        "Values": "'01/Truck 100 M3', '12/TC 45\" Dry 88 M3'",
    },
]

CARRIER_RULES = [
    {"Name": "Ares S.L.", "Operator": "CARRIER equals", "Values": "'ARES BILBAO S.L.'"},
    {"Name": "Containerships - CMA CGM GmbH", "Operator": "CARRIER equals", "Values": "'CMA CGM.SA'"},
    {"Name": "Dolt Logistics", "Operator": "CARRIER equals", "Values": "'DLS'"},
    {"Name": "ECS", "Operator": "CARRIER equals", "Values": "'ECS European Containers'"},
    {"Name": "Ewals Cargo Care (R5)", "Operator": "CARRIER equals", "Values": "'Ewals'"},
    {
        "Name": "FERCAM SPA (R2)",
        "Operator": "CARRIER equals",
        "Values": "'FERCAM SPA', 'Fercam BG', 'Fercam AT'",
    },
    {
        "Name": "G.T.S. SPA - General Transport Service",
        "Operator": "CARRIER equals",
        "Values": "'GTS TRASPORTI SPA'",
    },
    {"Name": "Gravis transport", "Operator": "CARRIER equals", "Values": "'Gravis'"},
    {"Name": "IntegreTrans", "Operator": "CARRIER equals", "Values": "'INTEGRE TRANS'"},
    {"Name": "J&S Speed Kft (R2)", "Operator": "CARRIER equals", "Values": "'J & S Speed Kft.'"},
    {"Name": "Marcotran", "Operator": "CARRIER equals", "Values": "'MARCOTRAN TISA'"},
    {"Name": "Miratrans", "Operator": "CARRIER equals", "Values": "'Miratrans Transport'"},
    {"Name": "O.K. Trans", "Operator": "CARRIER equals", "Values": "'O.K.TRANS PRAHA spol. s.r.o.'"},
    {"Name": "Patinter", "Operator": "CARRIER equals", "Values": "'PATINTER HQ'"},
    {"Name": "Patinter , SA (R6)", "Operator": "CARRIER equals", "Values": "'PATINTER HQ'"},
    {"Name": "PTM", "Operator": "CARRIER equals", "Values": "'PTM TRANSPORT sp. z o.o.'"},
    {
        "Name": "NINATRANS IBERIA LOGISTICA",
        "Operator": "CARRIER equals",
        "Values": "'NINATRANS IBERIA LOGISTICA SAU'",
    },
    {
        "Name": "SLC Transport - Szíjj Csaba (R2)",
        "Operator": "CARRIER equals",
        "Values": "'SLC Transport'",
    },
    {"Name": "Smet", "Operator": "CARRIER equals", "Values": "'SMET SPA'"},
    {"Name": "Stalspeed", "Operator": "CARRIER equals", "Values": "'STALSPEED Sp. z o.o.'"},
    {
        "Name": "STANTE LOGISTICS SRL",
        "Operator": "CARRIER equals",
        "Values": "'STANTE LOGISTICS SPA SB'",
    },
    {"Name": "Suus", "Operator": "CARRIER equals", "Values": "'SUUS Logistics SA'"},
    {
        "Name": "Szam",
        "Operator": "CARRIER equals",
        "Values": "'Szam Nemzetkozi Arufuvarozo Kft.'",
    },
    {"Name": "Traffic", "Operator": "CARRIER equals", "Values": "'Traffic Transport Sp.z o.o.'"},
    {"Name": "Trans Sped", "Operator": "CARRIER equals", "Values": "'Transsped'"},
    {"Name": "Transfennica", "Operator": "CARRIER equals", "Values": "'Transfennica logistics BVBA'"},
    {"Name": "Transintertop", "Operator": "CARRIER equals", "Values": "'Transintertop Kft'"},
    {
        "Name": "TRANSPORTES HNOS. LAREDO (R3)",
        "Operator": "CARRIER equals",
        "Values": "'TRANSPORTES HNOS. LAREDO S.A.'",
    },
    {
        "Name": "Vos Logistics Cargo",
        "Operator": "CARRIER equals",
        "Values": "'Vos Logistics Cargo International'",
    },
    {
        "Name": "XPO Logistics Spain",
        "Operator": "CARRIER equals",
        "Values": "'XPO TRANSPORT SOLUTIONS SPAIN SL'",
    },
    {
        "Name": "OnTime",
        "Operator": "CARRIER equals",
        "Values": "'ONTIME TRANSPORTE Y LOGISTICA SL'",
    },
    {
        "Name": "P&O Ferrymasters Limited (R2)",
        "Operator": "CARRIER equals",
        "Values": "'P&O Ferrymasters Ltd'",
    },
    {"Name": "Ravitex", "Operator": "CARRIER equals", "Values": "'RAVITEX SRL'"},
    {"Name": "DHL", "Operator": "CARRIER equals", "Values": "'DHL FREIGHT SPAIN SLU'"},
    {"Name": "CEVA Freight Poland", "Operator": "CARRIER equals", "Values": "'Ceva'"},
    {"Name": "Samskip", "Operator": "CARRIER equals", "Values": "'SAMSKIP MULTIMODAL B.V.'"},
    {"Name": "Chomar", "Operator": "CARRIER equals", "Values": "'CHOMAR Sp. z o.o.'"},
    {"Name": "Indeka/Indeka Sp. z o.o.", "Operator": "CARRIER equals", "Values": "'Indeka Sp. z o.o.'"},
    {"Name": "Logisteed Poland", "Operator": "CARRIER equals", "Values": "'LOGISTEED POLAND Sp. z o.o.'"},
    {"Name": "Baltic Transline", "Operator": "CARRIER equals", "Values": "'Baltic Transline UAB'"},
]


CONDITION_BLOCK_COLUMNS = {
    "Conditions for Equipment type": "Equipment type",
    "Conditions for Carrier": "Carrier",
    "Conditions for Destination city to differentiate lanes": "Destination city to differentiate lanes",
}

CITIES_TO_UPPERCASE = [
    "Astana",
    "Atessa",
    "Beaune",
    "Bischofsheim",
    "Bitonto",
    "Boblingen",
    "Bourne",
    "Bremen-hemelingen",
    "Bremerhaven",
    "Bucaresti",
    "Burgos",
    "Bydgoszcz",
    "Chanteloup Les Vignes",
    "Ditzingen",
    "Dublin",
    "Duisburg",
    "Dürrholz",
    "Ehingen",
    "Genk",
    "Gernsheim",
    "Ghent",
    "Giengen/Brenz",
    "Hainichen",
    "Holle",
    "Karaganda",
    "Kirchdorf",
    "Klingenberg",
    "Łowicz",
    "Paczkowo",
    "Philippsburg",
    "Sint-niklaas",
    "Stratford Upon Avon",
    "Stryków",
    "Tychy",
    "Wedemark",
    "Wilhelmshaven",
    "Poznan",
]


def _canonical_name(name: object) -> str:
    value = str(name).strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(char for char in value if not unicodedata.combining(char))
    for token in (" ", "_", "-", "/", "(", ")", '"', "'", ":", ".", ",", "#"):
        value = value.replace(token, "")
    return value


def list_extracted_files() -> list[Path]:
    if not PROCESSING_DIR.exists():
        return []
    return sorted(
        [p for p in PROCESSING_DIR.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS],
        key=lambda p: p.name.lower(),
    )


def choose_file(files: list[Path]) -> Path:
    print("Extracted files available in processing folder:\n")
    for idx, file_path in enumerate(files, start=1):
        print(f"{idx}. {file_path.name}")

    while True:
        user_input = input("\nChoose file number to format: ").strip()
        if not user_input.isdigit():
            print("Please enter a valid number.")
            continue
        selected_index = int(user_input)
        if 1 <= selected_index <= len(files):
            return files[selected_index - 1]
        print("Number is out of range. Try again.")


def load_dataframe(file_path: Path) -> pd.DataFrame:
    if file_path.suffix.lower() == ".csv":
        return pd.read_csv(file_path)
    return pd.read_excel(file_path)


def _find_bid_currency_index(df: pd.DataFrame) -> int | None:
    for idx, column in enumerate(df.columns):
        if _canonical_name(column) == "bidcurrency":
            return idx
    return None


def _extract_capacities(text: str) -> set[int]:
    text_l = text.lower()
    matches = re.findall(r"(\d+)\s*(?:m3|cubic\s*m|h-cube)", text_l)
    return {int(m) for m in matches}


def _rule_value_kind(rule: dict[str, str]) -> str:
    merged = f"{rule['Name']} {rule['Values']}".lower()
    if "container" in merged or "tc " in merged:
        return "container"
    return "truck"


def _normalize_equipment_type_values(df: pd.DataFrame) -> pd.DataFrame:
    equipment_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Equipment type")),
        None,
    )
    if equipment_col is None:
        return df

    direct_rule_by_name = {
        _canonical_name(rule["Name"]): rule["Name"] for rule in CONDITION_RULES
    }
    rule_capacities = {
        rule["Name"]: _extract_capacities(f"{rule['Name']} {rule['Values']}")
        for rule in CONDITION_RULES
    }
    rule_kind = {rule["Name"]: _rule_value_kind(rule) for rule in CONDITION_RULES}

    def _map_one(value: object) -> object:
        if pd.isna(value):
            return value

        raw = str(value).strip()
        if not raw:
            return value

        direct = direct_rule_by_name.get(_canonical_name(raw))
        if direct:
            return direct

        caps = _extract_capacities(raw)
        if not caps:
            return value

        value_kind = "container" if "container" in raw.lower() else "truck"
        candidates: list[tuple[int, str]] = []
        for rule in CONDITION_RULES:
            name = rule["Name"]
            rule_caps = rule_capacities.get(name, set())
            if not caps.issubset(rule_caps):
                continue
            if rule_kind.get(name) != value_kind:
                continue
            # Prefer tighter matches (fewer capacities in rule).
            candidates.append((len(rule_caps), name))

        if not candidates:
            return value
        candidates.sort(key=lambda item: item[0])
        return candidates[0][1]

    df = df.copy()
    df[equipment_col] = df[equipment_col].apply(_map_one)
    return df


def _strip_parentheses(text: str) -> str:
    return re.sub(r"\s*\([^)]*\)", "", text).strip()


def _split_csv_tokens(text: object) -> list[str]:
    if pd.isna(text):
        return []
    return [part.strip() for part in str(text).split(",") if part.strip()]


def _parse_condition_value_tokens(values_text: str) -> set[str]:
    quoted = re.findall(r"'([^']*)'", values_text)
    if quoted:
        return {_canonical_name(value) for value in quoted if value.strip()}
    return {_canonical_name(value) for value in _split_csv_tokens(values_text)}


def _equipment_condition_tokens(equipment_type: object) -> set[str]:
    if pd.isna(equipment_type):
        return set()

    canonical_equipment = _canonical_name(equipment_type)
    if not canonical_equipment:
        return set()

    tokens = {canonical_equipment}
    for rule in CONDITION_RULES:
        rule_name = _canonical_name(rule["Name"])
        rule_values = _parse_condition_value_tokens(rule["Values"])
        if canonical_equipment == rule_name or canonical_equipment in rule_values:
            tokens.add(rule_name)
            tokens.update(rule_values)
    return tokens


def _equipment_condition_values_intersect(equipment_a: object, equipment_b: object) -> bool:
    tokens_a = _equipment_condition_tokens(equipment_a)
    tokens_b = _equipment_condition_tokens(equipment_b)
    if not tokens_a or not tokens_b:
        return False
    return bool(tokens_a & tokens_b)


def _load_postal_code_zones(path: Path | None = None) -> dict[str, list[tuple[str, set[str]]]]:
    """Map canonical zone names to (country, postal-code tokens)."""
    zones_path = path or postal_code_zones_path()
    if not zones_path.exists():
        return {}

    zones_by_name: dict[str, list[tuple[str, set[str]]]] = {}
    raw = pd.read_csv(zones_path, sep="\t", dtype=str).fillna("")
    expected_columns = {"name", "country", "postal code"}
    if not expected_columns.issubset({_canonical_name(col) for col in raw.columns}):
        raw = pd.read_csv(
            zones_path,
            sep="\t",
            dtype=str,
            header=None,
            names=["Name", "Country", "Postal Code", "Excluded"],
        ).fillna("")
    for _, row in raw.iterrows():
        name_key = _canonical_name(row.get("Name", ""))
        if not name_key:
            continue

        country = str(row.get("Country", "")).strip().upper()
        postal_tokens = {
            _canonical_name(token)
            for token in _split_csv_tokens(row.get("Postal Code", ""))
            if str(token).strip()
        }
        excluded_tokens = {
            _canonical_name(token)
            for token in _split_csv_tokens(row.get("Excluded", ""))
            if str(token).strip()
        }
        if excluded_tokens:
            postal_tokens -= excluded_tokens
        if not country or not postal_tokens:
            continue

        zones_by_name.setdefault(name_key, []).append((country, postal_tokens))

    return zones_by_name


def _zone_keys_for_value(
    value: object,
    country_value: object,
    zones_by_name: dict[str, list[tuple[str, set[str]]]],
) -> set[tuple[str, str]]:
    if pd.isna(value):
        return set()

    text = str(value).strip()
    if not text:
        return set()

    keys: set[tuple[str, str]] = set()
    lookup_keys = [_canonical_name(text)]
    lookup_keys.extend(_canonical_name(token) for token in _split_csv_tokens(text))

    for lookup_key in lookup_keys:
        for country, postal_tokens in zones_by_name.get(lookup_key, []):
            for postal_token in postal_tokens:
                keys.add((country, postal_token))

    if keys:
        return keys

    country = str(country_value).strip().upper() if pd.notna(country_value) else ""
    fallback_token = _canonical_name(text)
    if country and fallback_token:
        return {(country, fallback_token)}
    if fallback_token:
        return {("", fallback_token)}
    return set()


def _zone_keys_intersect(keys_a: set[tuple[str, str]], keys_b: set[tuple[str, str]]) -> bool:
    return bool(keys_a & keys_b)


def _lanes_need_differentiation(
    row_a: pd.Series,
    row_b: pd.Series,
    *,
    carrier_col: str,
    equipment_col: str,
    origin_postal_col: str,
    origin_country_col: str,
    destination_col: str,
    destination_country_col: str,
    zones_by_name: dict[str, list[tuple[str, set[str]]]],
) -> bool:
    if str(row_a[carrier_col]).strip() != str(row_b[carrier_col]).strip():
        return False

    if not _equipment_condition_values_intersect(row_a[equipment_col], row_b[equipment_col]):
        return False

    origin_keys_a = _zone_keys_for_value(row_a[origin_postal_col], row_a[origin_country_col], zones_by_name)
    origin_keys_b = _zone_keys_for_value(row_b[origin_postal_col], row_b[origin_country_col], zones_by_name)
    destination_keys_a = _zone_keys_for_value(row_a[destination_col], row_a[destination_country_col], zones_by_name)
    destination_keys_b = _zone_keys_for_value(row_b[destination_col], row_b[destination_country_col], zones_by_name)

    origin_zone_overlap = _zone_keys_intersect(origin_keys_a, origin_keys_b)
    destination_zone_overlap = _zone_keys_intersect(destination_keys_a, destination_keys_b)
    if not (origin_zone_overlap and destination_zone_overlap):
        return False

    origin_values_differ = _canonical_name(row_a[origin_postal_col]) != _canonical_name(row_b[origin_postal_col])
    destination_values_differ = _canonical_name(row_a[destination_col]) != _canonical_name(
        row_b[destination_col]
    )
    return origin_values_differ or destination_values_differ


def _find_lane_differentiation_pairs(df: pd.DataFrame) -> list[tuple[object, object]]:
    """Return row-index pairs that need destination-city lane differentiation."""
    destination_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Destination City")),
        None,
    )
    equipment_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Equipment type")),
        None,
    )
    carrier_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Carrier")),
        None,
    )
    origin_postal_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Origin Postal Code")),
        None,
    )
    origin_country_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Origin Country")),
        None,
    )
    destination_country_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Destination Country")),
        None,
    )
    required_columns = (
        destination_col,
        equipment_col,
        carrier_col,
        origin_postal_col,
        origin_country_col,
        destination_country_col,
    )
    if any(column is None for column in required_columns):
        return []

    zones_by_name = _load_postal_code_zones()
    row_indexes = list(df.index)
    pairs: list[tuple[object, object]] = []
    for i in range(len(row_indexes)):
        row_a = df.loc[row_indexes[i]]
        for j in range(i + 1, len(row_indexes)):
            row_b = df.loc[row_indexes[j]]
            if _lanes_need_differentiation(
                row_a,
                row_b,
                carrier_col=carrier_col,
                equipment_col=equipment_col,
                origin_postal_col=origin_postal_col,
                origin_country_col=origin_country_col,
                destination_col=destination_col,
                destination_country_col=destination_country_col,
                zones_by_name=zones_by_name,
            ):
                pairs.append((row_indexes[i], row_indexes[j]))
    return pairs


def _normalized_shipment_cost(value: object) -> float | None:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return None
    return round(float(numeric), 2)


def _row_indexes_with_same_price_lane_pairs(df: pd.DataFrame) -> set[object]:
    """Rows in lane-differentiation pairs where both lanes share the same transport cost."""
    if df.empty or len(df.columns) == 0:
        return set()

    cost_col = df.columns[-1]
    highlighted: set[object] = set()
    for idx_a, idx_b in _find_lane_differentiation_pairs(df):
        cost_a = _normalized_shipment_cost(df.loc[idx_a, cost_col])
        cost_b = _normalized_shipment_cost(df.loc[idx_b, cost_col])
        if cost_a is None or cost_b is None:
            continue
        if cost_a == cost_b:
            highlighted.add(idx_a)
            highlighted.add(idx_b)
    return highlighted


def _cities_to_uppercase_keys() -> set[str]:
    return {_canonical_name(city) for city in CITIES_TO_UPPERCASE}


def _uppercase_listed_city(value: object) -> object:
    if pd.isna(value):
        return value

    text = str(value).strip()
    if not text:
        return value

    listed_keys = _cities_to_uppercase_keys()
    if _canonical_name(text) in listed_keys or _canonical_name(_strip_parentheses(text)) in listed_keys:
        return text.upper()
    return value


def _uppercase_listed_city_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for target_name in ("Destination City", "Destination city to differentiate lanes"):
        column = next(
            (col for col in df.columns if _canonical_name(col) == _canonical_name(target_name)),
            None,
        )
        if column is not None:
            df[column] = df[column].apply(_uppercase_listed_city)
    return df


def _fill_destination_city_to_differentiate_lanes(df: pd.DataFrame) -> pd.DataFrame:
    """Fill differentiate-lanes for fully duplicated lane pairs.

    A pair is filled only when all four dimensions intersect at the same time:
    Carrier, Equipment type (via condition rules), Origin Postal Code zones,
    and Destination City zones. At least one of the compared origin postal code
    or destination city raw values must differ.
    """
    destination_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Destination City")),
        None,
    )
    equipment_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Equipment type")),
        None,
    )
    carrier_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Carrier")),
        None,
    )
    origin_postal_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Origin Postal Code")),
        None,
    )
    origin_country_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Origin Country")),
        None,
    )
    destination_country_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Destination Country")),
        None,
    )
    diff_col = next(
        (
            col
            for col in df.columns
            if _canonical_name(col) == _canonical_name("Destination city to differentiate lanes")
        ),
        None,
    )
    required_columns = (
        destination_col,
        equipment_col,
        carrier_col,
        origin_postal_col,
        origin_country_col,
        destination_country_col,
        diff_col,
    )
    if any(column is None for column in required_columns):
        return df

    df = df.copy()
    df[diff_col] = df[diff_col].astype("object")

    rows_to_fill: set[object] = set()
    for idx_a, idx_b in _find_lane_differentiation_pairs(df):
        rows_to_fill.add(idx_a)
        rows_to_fill.add(idx_b)

    if rows_to_fill:
        df.loc[list(rows_to_fill), diff_col] = df.loc[list(rows_to_fill), destination_col]

    return df


def _round_transport_cost_to_2_decimals(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or len(df.columns) == 0:
        return df

    df = df.copy()
    cost_col = df.columns[-1]
    numeric_cost = pd.to_numeric(df[cost_col], errors="coerce")
    rounded_cost = numeric_cost.round(2)
    df.loc[numeric_cost.notna(), cost_col] = rounded_cost[numeric_cost.notna()]
    return df


def _normalize_carrier_values(df: pd.DataFrame) -> pd.DataFrame:
    carrier_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Carrier")),
        None,
    )
    if carrier_col is None:
        return df

    df = df.copy()
    df[carrier_col] = df[carrier_col].replace({"Ellerman Lines": "Ellerman"})
    return df


def _build_conditions_blocks(df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    equipment_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Equipment type")),
        None,
    )
    carrier_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Carrier")),
        None,
    )
    destination_col = next(
        (col for col in df.columns if _canonical_name(col) == _canonical_name("Destination City")),
        None,
    )
    diff_col = next(
        (
            col
            for col in df.columns
            if _canonical_name(col) == _canonical_name("Destination city to differentiate lanes")
        ),
        None,
    )
    equipment_rows: list[dict[str, object]] = []
    if equipment_col is not None:
        equipment_values = {
            _canonical_name(value)
            for value in df[equipment_col].dropna().astype(str).tolist()
            if str(value).strip()
        }

        matched_rules = [
            rule
            for rule in CONDITION_RULES
            if _canonical_name(rule["Name"]) in equipment_values
        ]

        equipment_rows = [
            {
                "Rule #": idx,
                "Name": rule["Name"],
                "Operator": rule["Operator"],
                "Values": rule["Values"],
                "Scope": "",
            }
            for idx, rule in enumerate(matched_rules, start=1)
        ]

    city_rows: list[dict[str, object]] = []
    if destination_col is not None and diff_col is not None:
        city_values = (
            df.loc[df[diff_col].notna(), destination_col]
            .dropna()
            .astype(str)
            .map(str.strip)
        )
        for city in sorted(set(v for v in city_values.tolist() if v)):
            city_rows.append(
                {
                    "Rule #": len(city_rows) + 1,
                    "Name": city,
                    "Operator": "starts with",
                    "Values": _strip_parentheses(city).upper(),
                    "Scope": "all items",
                }
            )

    blocks: list[tuple[str, pd.DataFrame]] = []
    if equipment_rows:
        blocks.append(
            (
                "Conditions for Equipment type",
                pd.DataFrame(equipment_rows, columns=["Rule #", "Name", "Operator", "Values", "Scope"]),
            )
        )

    if carrier_col is not None:
        carrier_values = {
            _canonical_name(value)
            for value in df[carrier_col].dropna().astype(str).tolist()
            if str(value).strip()
        }
        matched_carrier_rules = [
            rule
            for rule in CARRIER_RULES
            if _canonical_name(rule["Name"]) in carrier_values
        ]
        if matched_carrier_rules:
            carrier_rows = [
                {
                    "Rule #": idx,
                    "Name": rule["Name"],
                    "Operator": rule["Operator"],
                    "Values": rule["Values"],
                    "Scope": "",
                }
                for idx, rule in enumerate(matched_carrier_rules, start=1)
            ]
            blocks.append(
                (
                    "Conditions for Carrier",
                    pd.DataFrame(
                        carrier_rows,
                        columns=["Rule #", "Name", "Operator", "Values", "Scope"],
                    ),
                )
            )

    if city_rows:
        blocks.append(
            (
                "Conditions for Destination city to differentiate lanes",
                pd.DataFrame(city_rows, columns=["Rule #", "Name", "Operator", "Values", "Scope"]),
            )
        )
    return blocks


def save_formatted_output(df: pd.DataFrame, source_file: Path) -> Path:
    if df.empty or len(df.columns) == 0:
        raise ValueError("Source file is empty.")

    bid_currency_idx = _find_bid_currency_index(df)
    if bid_currency_idx is None:
        raise ValueError("Column 'Bid Currency' was not found in the source file.")

    last_col_idx = len(df.columns) - 1
    if bid_currency_idx > last_col_idx:
        raise ValueError("Invalid source layout.")

    df = _normalize_equipment_type_values(df)
    df = _normalize_carrier_values(df)
    df = _fill_destination_city_to_differentiate_lanes(df)
    df = _uppercase_listed_city_columns(df)
    df = _round_transport_cost_to_2_decimals(df)
    same_price_lane_rows = _row_indexes_with_same_price_lane_pairs(df)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_DIR / f"{source_file.stem}_output.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Keep original column names and order exactly as in source.
        df.to_excel(writer, sheet_name="Formatted", startrow=3, index=False)
        ws = writer.sheets["Formatted"]

        route_fill = PatternFill("solid", fgColor="DCE6F1")
        rates_fill = PatternFill("solid", fgColor="E2F0D9")
        same_price_lane_fill = PatternFill("solid", fgColor="DCE6F1")
        header_fill = PatternFill("solid", fgColor="F2F2F2")
        bold_font = Font(bold=True)
        regular_font = Font(bold=False)
        center = Alignment(horizontal="center", vertical="center")

        # Top grouped layout row, similar to provided template structure.
        route_end_col = bid_currency_idx
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=route_end_col)
        route_cell = ws.cell(row=1, column=1, value=None)
        route_cell.font = regular_font
        route_cell.alignment = center
        route_cell.fill = route_fill

        ws.merge_cells(
            start_row=1,
            start_column=bid_currency_idx + 1,
            end_row=1,
            end_column=last_col_idx + 1,
        )
        rates_cell = ws.cell(row=1, column=bid_currency_idx + 1, value="Transport cost")
        rates_cell.font = bold_font
        rates_cell.alignment = center
        rates_cell.fill = rates_fill

        ws.merge_cells(
            start_row=2,
            start_column=bid_currency_idx + 1,
            end_row=2,
            end_column=last_col_idx + 1,
        )
        apply_if_cell = ws.cell(row=2, column=bid_currency_idx + 1, value="Apply if")
        apply_if_cell.font = regular_font
        apply_if_cell.alignment = center
        apply_if_cell.fill = rates_fill

        ws.merge_cells(
            start_row=3,
            start_column=bid_currency_idx + 1,
            end_row=3,
            end_column=last_col_idx + 1,
        )
        rate_by_cell = ws.cell(row=3, column=bid_currency_idx + 1, value="Rate by: Per shipment")
        rate_by_cell.font = regular_font
        rate_by_cell.alignment = center
        rate_by_cell.fill = rates_fill

        # Style/display the row with source column names (with requested label overrides).
        only_bold_headers = {
            _canonical_name("Origin Postal Code"),
            _canonical_name("Destination City"),
            _canonical_name("Destination city to differentiate lanes"),
            _canonical_name("Equipment type"),
            _canonical_name("Flat"),
        }
        for col_idx in range(1, last_col_idx + 2):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = regular_font
            cell.fill = header_fill
            if col_idx - 1 == bid_currency_idx:
                cell.value = "Currency"
            elif col_idx - 1 == last_col_idx:
                cell.value = "Flat"

            if _canonical_name(cell.value) in only_bold_headers:
                cell.font = bold_font

        ws.freeze_panes = "A5"

        conditions_blocks = _build_conditions_blocks(df)
        condition_names_by_column: dict[int, set[str]] = {}
        if not conditions_blocks:
            pd.DataFrame(columns=["Rule #", "Name", "Operator", "Values", "Scope"]).to_excel(
                writer,
                sheet_name="conditions",
                index=False,
            )
        else:
            current_row = 0
            for block_title, block_df in conditions_blocks:
                pd.DataFrame([[block_title]]).to_excel(
                    writer,
                    sheet_name="conditions",
                    startrow=current_row,
                    index=False,
                    header=False,
                )
                current_row += 1
                block_df.to_excel(
                    writer,
                    sheet_name="conditions",
                    startrow=current_row,
                    index=False,
                )
                target_column = CONDITION_BLOCK_COLUMNS.get(block_title)
                if target_column is not None:
                    col_idx = next(
                        (
                            idx + 1
                            for idx, col_name in enumerate(df.columns)
                            if _canonical_name(col_name) == _canonical_name(target_column)
                        ),
                        None,
                    )
                    if col_idx is not None:
                        block_names = {
                            str(v).strip()
                            for v in block_df.get("Name", pd.Series(dtype="object")).dropna().tolist()
                            if str(v).strip()
                        }
                        condition_names_by_column.setdefault(col_idx, set()).update(block_names)
                current_row += len(block_df) + 2

        if condition_names_by_column:
            highlight_fill = PatternFill("solid", fgColor="D9D9D9")
            data_start_row = 5
            data_end_row = ws.max_row
            data_end_col = ws.max_column
            destination_city_col = next(
                (
                    idx + 1
                    for idx, col_name in enumerate(df.columns)
                    if _canonical_name(col_name) == _canonical_name("Destination City")
                ),
                None,
            )
            origin_location_col = next(
                (
                    idx + 1
                    for idx, col_name in enumerate(df.columns)
                    if _canonical_name(col_name) == _canonical_name("ORIGIN LOCATION")
                ),
                None,
            )

            for row_idx in range(data_start_row, data_end_row + 1):
                for col_idx in range(1, data_end_col + 1):
                    if (
                        (destination_city_col is not None and col_idx == destination_city_col)
                        or (origin_location_col is not None and col_idx == origin_location_col)
                    ):
                        continue
                    allowed_names = condition_names_by_column.get(col_idx)
                    if not allowed_names:
                        continue
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is None:
                        continue
                    cell_text = str(cell.value).strip()
                    if cell_text in allowed_names:
                        cell.fill = highlight_fill
                        cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            color=cell.font.color,
                            underline="single",
                        )

        if same_price_lane_rows:
            data_end_col = ws.max_column
            index_to_excel_row = {idx: 5 + pos for pos, idx in enumerate(df.index)}
            for row_index in same_price_lane_rows:
                excel_row = index_to_excel_row.get(row_index)
                if excel_row is None:
                    continue
                for col_idx in range(1, data_end_col + 1):
                    ws.cell(row=excel_row, column=col_idx).fill = same_price_lane_fill

    return output_file


def main() -> int:
    files = list_extracted_files()
    if not files:
        print("No extracted files found in processing folder.")
        return 1

    selected_file = choose_file(files)
    print(f"\nFormatting '{selected_file.name}'...")
    df = load_dataframe(selected_file)
    output_path = save_formatted_output(df, selected_file)
    print(f"Done. Saved formatted output to '{output_path}'.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nOperation cancelled.")
        raise SystemExit(130)
    except Exception as exc:
        print(f"Error: {exc}")
        raise SystemExit(1)
